#!/usr/bin/env python3
"""
Codon optimization FASTA pipeline
  1. Merge one FASTA per subfolder  →  merged.fasta
  2. Reverse-complement all seqs    →  merged_revcomp.fasta
  3. Export name + revcomp seq      →  sequences_revcomp.xlsx

Usage:
    python fasta_pipeline.py <input_folder> [--pick first|last|longest]

Arguments:
    input_folder   Root folder produced by the codon-opt tool
                   (contains one sub-folder per sequence)
    --pick         Which FASTA to select when a sub-folder has multiple files
                   first   (default) – alphabetically first
                   last    – alphabetically last
                   longest – file with the most sequence characters
"""

import os
import re
import re
import sys
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


# ── FASTA helpers ─────────────────────────────────────────────────────────────

def parse_fasta(path: Path) -> list[tuple[str, str]]:
    """Return list of (header, sequence) tuples from a FASTA file."""
    records = []
    header, seq_parts = None, []
    with open(path) as fh:
        for line in fh:
            line = line.rstrip()
            if line.startswith(">"):
                if header is not None:
                    records.append((header, "".join(seq_parts)))
                header, seq_parts = line[1:], []
            else:
                seq_parts.append(line)
    if header is not None:
        records.append((header, "".join(seq_parts)))
    return records


def reverse_complement(seq: str) -> str:
    comp = str.maketrans("ACGTacgtNn", "TGCAtgcaNn")
    return seq.translate(comp)[::-1]


def pick_fasta(fasta_files: list[Path], strategy: str) -> Path:
    if strategy == "last":
        return sorted(fasta_files)[-1]
    if strategy == "longest":
        return max(fasta_files, key=lambda p: sum(
            len(l.rstrip()) for l in open(p) if not l.startswith(">")))
    return sorted(fasta_files)[0]   # "first"


# ── Step 1 – merge ────────────────────────────────────────────────────────────

def merge_fastas(root: Path, strategy: str, out_path: Path) -> list[tuple[str, str]]:
    extensions = {".fa", ".fasta", ".fna", ".ffn", ".frn"}
    all_records = []
    subfolders = sorted([d for d in root.iterdir() if d.is_dir()])

    if not subfolders:
        print(f"  [warn] No sub-folders found in {root}. "
              "Trying root folder directly.")
        subfolders = [root]

    for folder in subfolders:
        fastas = [f for f in folder.iterdir()
                  if f.is_file() and f.suffix.lower() in extensions]
        if not fastas:
            print(f"  [skip] {folder.name} – no FASTA files found")
            continue
        chosen = pick_fasta(fastas, strategy)
        records = parse_fasta(chosen)
        if not records:
            print(f"  [skip] {chosen} – empty or unparseable")
            continue
        # Strip trailing _<number> added by the codon-opt tool
        clean_name = re.sub(r"_\d+$", "", folder.name)
        records = [(clean_name, seq) for _, seq in records]
        print(f"  [merge] {folder.name}  →  '{clean_name}'  ←  {chosen.name}")
        all_records.extend(records)

    with open(out_path, "w") as fh:
        for header, seq in all_records:
            fh.write(f">{header}\n")
            for i in range(0, len(seq), 80):
                fh.write(seq[i:i+80] + "\n")

    print(f"\n✓ Merged {len(all_records)} sequences  →  {out_path}")
    return all_records


# ── Step 2 – reverse complement ───────────────────────────────────────────────

def make_revcomp_fasta(records: list[tuple[str, str]], out_path: Path) -> list[tuple[str, str]]:
    rc_records = [(hdr, reverse_complement(seq)) for hdr, seq in records]
    with open(out_path, "w") as fh:
        for header, seq in rc_records:
            fh.write(f">{header}\n")
            for i in range(0, len(seq), 80):
                fh.write(seq[i:i+80] + "\n")
    print(f"✓ Reverse-complemented {len(rc_records)} sequences  →  {out_path}")
    return rc_records


# ── Step 3 – Excel export ─────────────────────────────────────────────────────

def write_excel(rc_records: list[tuple[str, str]], out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "RevComp Sequences"

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    cell_font   = Font(name="Arial", size=10)
    wrap_align  = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 80

    for col, label in enumerate(["Sequence Name", "Reverse Complement Sequence"], start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 18

    for row_idx, (header, seq) in enumerate(rc_records, start=2):
        name_cell = ws.cell(row=row_idx, column=1, value=header)
        seq_cell  = ws.cell(row=row_idx, column=2, value=seq)
        for cell in (name_cell, seq_cell):
            cell.font = cell_font
            cell.alignment = wrap_align

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(out_path)
    print(f"✓ Excel file written ({len(rc_records)} rows)  →  {out_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("input_folder", help="Root output folder from codon-opt tool")
    parser.add_argument("--pick", choices=["first", "last", "longest"],
                        default="first",
                        help="How to pick one FASTA per sub-folder (default: first)")
    args = parser.parse_args()

    root = Path(args.input_folder).resolve()
    if not root.is_dir():
        sys.exit(f"Error: {root} is not a directory.")

    out_dir = root.parent   # place outputs next to the input folder
    merged_fasta  = out_dir / "merged.fasta"
    revcomp_fasta = out_dir / "merged_revcomp.fasta"
    excel_out     = out_dir / "sequences_revcomp.xlsx"

    print(f"\n── Step 1: Merging FASTAs from {root} ──")
    records = merge_fastas(root, args.pick, merged_fasta)

    if not records:
        sys.exit("No sequences found – nothing to do.")

    print(f"\n── Step 2: Reverse complementing ──")
    rc_records = make_revcomp_fasta(records, revcomp_fasta)

    print(f"\n── Step 3: Writing Excel ──")
    write_excel(rc_records, excel_out)

    print(f"""
┌─────────────────────────────────────────────────────────┐
│  Done! Output files:                                    │
│    merged.fasta              – all selected sequences   │
│    merged_revcomp.fasta      – reverse complement FASTA │
│    sequences_revcomp.xlsx    – name + revcomp for Excel │
└─────────────────────────────────────────────────────────┘
""")


if __name__ == "__main__":
    main()
